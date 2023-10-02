# this imports all of the libraries being used in the program
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
import pyautogui as pag
import os

# This tells the program where the In Process file is located
path = r'C:\Users\SHARON\Desktop\Deposits\LBVWS\In Process.xlsx'
# This loads the workbook into the program and data_only means that no formulas are being copied over
wb = load_workbook(path, data_only=True)
# this loads in and defines "Sheet1" Sheet
ws = wb['Sheet1']

# This finds and starts the UB program
os.startfile(r'C:\Users\Public\Desktop\WS Utility Billing')
# time.sleep is telling the program to wait the specified number of seconds
time.sleep(3)

# if you look up in the import, we imported pyautogui as pag
# here we are telling the computer to hit the enter button
pag.press('enter')
time.sleep(1)

# as you probably guessed from the previous one, this is telling the computer to hit the alt button
pag.press(['alt'])
# this time it is telling the computer to hit the right arrow 3 times
pag.press('right', presses=3)
# like last time except down 16 times
pag.press('down', presses=16)
pag.press('enter')
time.sleep(.5)

# This is telling the computer to click the mouse button at (1320,300) which is on the coordinate grid of your computer
pag.click(1320,300)
time.sleep(.1)

# this is our indexing variable and it is going to be used to determine which cell we are working with in the
ini = 3

# this is the beginning of a while loop
# essentially the code will keep running over and over again until the while statement is no longer true
# In this case the statement that it will be evaluating is the data type of the cell being equal to "s" (I have no clue what s means)
while ws["A"+str(ini)].data_type == "s":
    # this is grabbing the account number out of the A collumn
    acct = str(ws["A"+str(ini)].value)
    pag.typewrite(acct)
    # This is typing it into the UB
    time.sleep(.2)
    pag.press('tab')
    # this is one getting the value for the amount
    amt = str(ws["C"+str(ini)].value)
    pag.typewrite(amt)
    time.sleep(.2)
    pag.press('enter')
    time.sleep(.2)
    # this is our indexing variable getting increased so we can go down to the next cell
    ini = int(ini) + 1

print("done")
