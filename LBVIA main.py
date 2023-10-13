# This imports all of the libraries being used in the program
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
import pyautogui as pag
import os

# this path variable is where your excel sheet that you loaded the checks on is located
path = r'C:\Users\SHARON\Desktop\Deposits\LBVIA\In Progress.xlsx'
# wb is loading in the workbook to the program and is only giving the data no formulas or fomats
wb = load_workbook(path, data_only=True)
# ws is telling the program what worksheet
ws = wb['Sheet1']

# this is is telling the computer to start up the UB application 
os.startfile(r'C:\Users\SHARON\Desktop\LBVIA UB')
# time.sleep is telling the program to wait a 3 seconds 
time.sleep(3)

# pag.press is telling the computer to press enter
pag.press('enter')
time.sleep(1)

pag.press(['alt'])
# in this case it is telling the computer to have 3 right presses
pag.press('right', presses=3)
pag.press('down', presses=16)
pag.press('enter')
time.sleep(.5)

pag.click(1320,300)
time.sleep(.1)
# ini is a variable I created to tell the computer what row to look on
ini = 3


# this is a while loop and it is telling the program to keep repeating this part as long as the data type of the cell is "s"
# what is "s", I have no clue
while ws["A"+str(ini)].data_type == "s":
    # this finds the account number in excel
    acct = str(ws["A"+str(ini)].value)
    # this writes the account number in UB
    pag.typewrite(acct)
    time.sleep(.2)
    pag.press('tab')
    # this finds the amount in excel
    amt = str(ws["C"+str(ini)].value)
    pag.typewrite(amt)
    time.sleep(.2)
    pag.press('enter')
    time.sleep(.2)
    ini = int(ini) + 1

print("done")
